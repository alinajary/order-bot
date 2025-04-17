from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class csv2excel:
    """
    A class to convert CSV files to Excel files with formatting.
    """

    def __init__(self, csv_file: str, excel_file: str):
        """
        Initialize the csv2excel object.

        :param csv_file: Path to the input CSV file.
        :param excel_file: Path to the output Excel file.
        """
        self.csv_file = csv_file
        self.excel_file = excel_file

    def convert(self):
        """
        Convert the CSV file to an Excel file with formatting.
        """
        import pandas as pd

        # Read the CSV file
        df = pd.read_csv(self.csv_file)

        # Write to an Excel file
        df.to_excel(self.excel_file, index=False)

        # Apply formatting
        self._apply_formatting()

    def _apply_formatting(self):
        """
        Apply formatting to the Excel file.
        """
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = load_workbook(self.excel_file)
        ws = wb.active

        # Apply header formatting
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Light gray background
        header_font = Font(bold=True, color="000000")  # Black text
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:  # First row (header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        state_col_idx = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value.lower() == "state":
                state_col_idx = idx
                break

        state_colors = {
        "pending": "FFFACD",    # LemonChiffon
        "approved": "C6EFCE",   # Light green
        "rejected": "FFC7CE",   # Light red
        "delivered": "CFE2F3",  # Light blue
        "cancelled": "F4CCCC",  # Light pink
        # Add more states and colors as needed
        }
        if state_col_idx is not None:
            for row in ws.iter_rows(min_row=2, min_col=state_col_idx, max_col=state_col_idx):
                cell = row[0]
                state_value = str(cell.value).lower() if cell.value else ""
                color = state_colors.get(state_value)
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        # Apply alignment to all cells (including data rows)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Draw borders for all cells
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save the formatted Excel file
        wb.save(self.excel_file)